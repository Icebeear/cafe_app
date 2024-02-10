import json

import openpyxl
import pandas as pd
import requests  # type: ignore
from fastapi.encoders import jsonable_encoder

from src.redis.utils import redis


def google_task_local() -> None:

    r = redis.get_redis_client()

    file_path = 'admin/Menu.xlsx'

    df = pd.read_excel(file_path, header=None)

    all_data = [list(row[:7]) for index, row in df.iterrows()]

    res = []

    discounts: dict[str, str] = {}

    for table, row in enumerate(all_data):

        if str(row[0]) != 'nan':
            res.append(
                {
                    'id': row[0],
                    'table': table,
                    'title': row[1],
                    'description': row[2],
                    'submenus': []
                }
            )

        elif str(row[1]) != 'nan':
            res[-1]['submenus'].append(
                {
                    'id': row[1],
                    'table': table,
                    'title': row[2],
                    'description': row[3],
                    'dishes': []
                }
            )

        elif str(row[2]) != 'nan':

            try:
                discount = str(row[6])
            except IndexError:
                discount = 'nan'

            res[-1]['submenus'][-1]['dishes'].append(
                {
                    'id': row[2],
                    'table': table,
                    'title': row[3],
                    'description': row[4],
                    'price': str(row[5]),
                    'discount': discount,
                }
            )

    base_url = 'http://127.0.0.1:8000/api/v1'

    url = f'{base_url}/menus/nested'

    response = requests.get(url)

    wb = openpyxl.load_workbook('admin/Menu.xlsx')
    sheet = wb.active

    dishes = requests.get(f'{base_url}/menus/menu_id/submenus/google_submenu_id/dishes/')
    dishes_ids = [dish['id'] for dish in dishes.json()] or []

    menus = requests.get(f'{base_url}/menus/')
    menus_ids = [menu['id'] for menu in menus.json()] or []

    google_dishes_ids: list[str] = []

    def update_all_data(menu_id: str) -> None:

        response = requests.get(f'{base_url}/menus/{menu_id}/submenus/')

        submenus_ids = [submenu['id'] for submenu in response.json()] or []
        submenus_titles = [submenu['title'] for submenu in response.json()] or []

        google_submenus_ids = []

        for google_submenu in google_menu['submenus']:
            google_submenu_id = google_submenu['id']
            google_submenu_title = google_submenu['title']

            if google_submenu_id in submenus_ids:
                requests.patch(f'{base_url}/menus/{menu_id}/submenus/{google_submenu_id}', json={
                    'title': google_submenu['title'],
                    'description': google_submenu['description'],
                })

            elif google_submenu_title in submenus_titles:

                submenu_id = [submenu['id'] for submenu in submenus_ids if submenu['title'] == google_submenu_title][0]

                requests.patch(f'{base_url}/menus/{menu_id}/submenus/{submenu_id}', json={
                    'title': google_submenu['title'],
                    'description': google_submenu['description'],
                })

                google_submenu_id = submenu_id

            else:
                response = requests.post(f'{base_url}/menus/{menu_id}/submenus/', json={
                    'title': google_submenu['title'],
                    'description': google_submenu['description'],
                })

                google_submenu_id = response.json()['id']

            sheet.cell(row=google_submenu['table'] + 1, column=2).value = google_submenu_id

            for google_dish in google_submenu['dishes']:
                google_dish_id = google_dish['id']

                if google_dish_id in dishes_ids:
                    requests.patch(f'{base_url}/menus/{menu_id}/submenus/{google_submenu_id}/dishes/{google_dish_id}', json={
                        'title': google_dish['title'],
                        'description': google_dish['description'],
                        'price': google_dish['price'],
                    })

                else:
                    response = requests.post(f'{base_url}/menus/{menu_id}/submenus/{google_submenu_id}/dishes/', json={
                        'title': google_dish['title'],
                        'description': google_dish['description'],
                        'price': google_dish['price'],
                    })

                    google_dish_id = response.json()['id']

                    sheet.cell(row=google_dish['table'] + 1, column=3).value = google_dish_id

                google_dishes_ids.append(google_dish_id)

                discounts[google_dish_id] = google_dish['discount']

            google_submenus_ids.append(google_submenu_id)

        # удаляем подменю которые есть в бд но нет в таблице
        for submenu_id in submenus_ids:
            if submenu_id not in google_submenus_ids:
                requests.delete(f'{base_url}/menus/{menu_id}/submenus/{submenu_id}')

    def create_all_data(menu_id: str) -> None:

        sheet.cell(row=google_menu['table'] + 1, column=1).value = menu_id

        for google_submenu in google_menu['submenus']:
            response = requests.post(f'{base_url}/menus/{menu_id}/submenus/', json={
                'title': google_submenu['title'],
                'description': google_submenu['description'],
            })

            if response.status_code == 201:

                submenu_id = response.json()['id']
                sheet.cell(row=google_submenu['table'] + 1, column=2).value = submenu_id

                for google_dish in google_submenu['dishes']:
                    response = requests.post(f'{base_url}/menus/{menu_id}/submenus/{submenu_id}/dishes/', json={
                        'title': google_dish['title'],
                        'description': google_dish['description'],
                        'price': google_dish['price'],
                    })

                    dish_id = response.json()['id']
                    sheet.cell(row=google_dish['table'] + 1, column=3).value = dish_id

    google_menus_ids = []
    for google_menu in res:
        menu_id = google_menu['id']
        google_menus_ids.append(menu_id)

        response = requests.get(f'{base_url}/menus/{menu_id}')

        '''
        Идем в бд по id меню из таблицы, если оно есть, сразу патчим
        и обновляем зависимости.
        '''
        if response.status_code == 200:
            requests.patch(f'{base_url}/menus/{menu_id}', json={
                'title': google_menu['title'],
                'description': google_menu['description'],
            })

            update_all_data(menu_id)

        else:
            '''
            Если не нашли по id (значит меню либо нету, либо до этого когда то создавалось в бд вручную),
            пытаемся создать, если получилось,
            создаем все остальное из таблицы. Если нет, значит в бд уже лежит меню с таким названием.
            Находим его по названию, берем id и идем обновлять все зависимости.
            '''
            response = requests.post(f'{base_url}/menus/', json={
                'title': google_menu['title'],
                'description': google_menu['description'],
            })

            if response.status_code == 201:
                menu_id = response.json()['id']

                create_all_data(menu_id)

            elif response.status_code == 409:
                all_menus = requests.get(f'{base_url}/menus/')
                menu_id = [menu['id'] for menu in all_menus.json() if google_menu['title'] == menu['title']][0]

                update_all_data(menu_id)

    # удаляем блюда которые есть в бд но нет в таблице
    for dish_id in dishes_ids:
        if dish_id not in google_dishes_ids:
            requests.delete(f'{base_url}/menus/menu_id/submenus/submenu_id/dishes/{dish_id}')

    # удаляем меню которые есть в бд но нет в таблице
    for menu_id in menus_ids:
        if menu_id not in google_menus_ids:
            requests.delete(f'{base_url}/menus/{menu_id}')

    r.set('discounts', json.dumps(jsonable_encoder(discounts)))

    wb.save('admin/Menu.xlsx')
